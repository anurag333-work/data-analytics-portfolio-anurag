# -*- coding: utf-8 -*-
"""excel_automation_python.ipynb
#**Excel Automation Using Python**

**Task:** Automate Monthly Sales Report Generation from Multiple Excel Files

**Client Requirements:** Client runs an e-commerce store and every month he receives 10-20 Excel files from different regional managers.

Each file contains: Order ID, Date, Region, Product Category, Units Sold, Unit Price

As manually merging these files and generating insights takes a lot of time. Client needs a Python script that will:
1. Merge Data Automatically
2. Data Cleaning
3. Generate Automated Analysis
4. Deliver a Single File (Output): Automated_Sales_Report.xlsx
"""

#import libraries

import pandas as pd
import os
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
from datetime import datetime

input_dir = '../input_files'
output_file = '../Automated_Sales_Report.xlsx'

# Create workbook
output_wb = Workbook()
output_sheet = output_wb.active
output_sheet.title = 'Clean Combined Data'

dest_row = 1
destination_start_col = 1

# ---------------------------
# 1. MERGE EXCEL FILES
# ---------------------------
for filename in os.listdir(input_dir):
    if filename.endswith('.xlsx'):
        source_path = os.path.join(input_dir, filename)
        print("Working on:", filename)

        source_wb = openpyxl.load_workbook(source_path)
        source_ws = source_wb.active

        # Determine source start row
        source_start_row = 1 if dest_row == 1 else 2
        source_end_row = source_ws.max_row
        source_end_col = source_ws.max_column

        # Copy values
        for row in range(source_start_row, source_end_row + 1):
            for col in range(1, source_end_col + 1):

                dest_col = destination_start_col + (col - 1)
                value = source_ws.cell(row=row, column=col).value

                # Fix date format (skip header row)
                if dest_row > 1 and col == 2:
                    value = datetime.strptime(value, "%Y-%m-%d").date()

                # Fix inconsistent region names
                if dest_row > 1 and col == 3:
                    value = str(value).title().strip()

                output_sheet.cell(row=dest_row, column=dest_col).value = value

                # Add file name column
                if col == source_end_col:
                    if dest_row == 1:
                        output_sheet.cell(row=dest_row, column=dest_col + 1).value = "File Name"
                    else:
                        output_sheet.cell(row=dest_row, column=dest_col + 1).value = filename

            dest_row += 1

# Header formatting
for col in range(1, output_sheet.max_column + 1):
    output_sheet[f"{get_column_letter(col)}1"].font = Font(bold=True)

# Save merged file
output_wb.save(output_file)

print("Merging complete. Working on summary...")

# ---------------------------
# 2. READ MERGED DATAFRAME
# ---------------------------
df = pd.read_excel(output_file)

df["Date"] = pd.to_datetime(df["Date"])
df["Month"] = df["Date"].dt.to_period("M")
df["Region"] = df["Region"].str.title().str.strip()
df["Revenue"] = df["Units Sold"] * df["Unit Price"]

# ---------------------------
# 3. MONTHLY SUMMARY
# ---------------------------
monthly_summary = (
    df.groupby("Month")
      .agg(
          Total_Units=("Units Sold", "sum"),
          Total_Revenue=("Revenue", "sum"),
          Total_Orders=("Order_ID", "count")
      )
      .assign(AOV=lambda x: x["Total_Revenue"] / x["Total_Orders"])
      .reset_index()
)

# ---------------------------
# 4. REGION SUMMARY
# ---------------------------
regional_summary = (
    df.groupby("Region")
      .agg(
          Units_Sold=("Units Sold", "sum"),
          Revenue=("Revenue", "sum")
      )
      .reset_index()
)

# ---------------------------
# 5. WRITE ALL SUMMARIES IN ONE GO
# ---------------------------
with pd.ExcelWriter(output_file, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
    monthly_summary.to_excel(writer, sheet_name="Monthly Summary", index=False)
    regional_summary.to_excel(writer, sheet_name="Region-wise Summary", index=False)

print("Process Complete")
